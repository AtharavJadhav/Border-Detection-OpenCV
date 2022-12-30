import cv2
import numpy as np
import os
from os.path import join
from os import listdir
from PIL import Image, ImageDraw, ImageFilter
from rembg import remove
import openpyxl
import numpy as np

path3 = 'C:/Users/Atharav Jadhav/source/repos/OpenCV project/Final Images/'
path2 = 'C:/Users/Atharav Jadhav/source/repos/OpenCV project/PNG Images/'
path1 = 'C:/Users/Atharav Jadhav/source/repos/OpenCV project/Cropped Images/'
folder = 'C:/Users/Atharav Jadhav/source/repos/OpenCV project/Trial Images/'

count = 1
def image_resize(image, width = None, height = None, inter = cv2.INTER_AREA):
    dim = None
    (h, w) = image.shape[:2]

    if width is None and height is None:
        return image
    if width is None:
        r = height / float(h)
        dim = (int(w * r), height)
    else:
        r = width / float(w)
        dim = (width, int(h * r))

    resized = cv2.resize(image, dim, interpolation = inter)

    return resized

fname = "Coordinates.xlsx"
wb = openpyxl.load_workbook(fname)
ws = wb['Sheet1']

for filename in os.listdir(folder):
       img = cv2.imread(os.path.join(folder,filename))
       if img is not None: 
           img = image_resize(img, height = 600)
           roi = cv2.selectROI(img)
           (x1, y1, width, height) = roi

           ws.cell(1,count).value = x1
           ws.cell(2,count).value = y1
           ws.cell(3,count).value = width
           ws.cell(4,count).value = height

           #print(type(roi))
           im_cropped = img[int(roi[1]):int(roi[1]+roi[3]),int(roi[0]):int(roi[0]+roi[2])]
           rs = str(count)
           image = remove(im_cropped)

           gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
           blurred = cv2.GaussianBlur(gray, (3, 3), 0)
           edged = cv2.Canny(blurred, 10, 100)

           # define a (3, 3) structuring element
           kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))

           # apply the dilation operation to the edged image
           dilate = cv2.dilate(edged, kernel, iterations=1)

           # find the contours in the dilated image
           contours, _ = cv2.findContours(dilate, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
           image_copy = image.copy()
           # draw the contours on a copy of the original image
           cv2.drawContours(image_copy, contours, -1, (0, 255, 0), 2)
           #print(len(contours), "objects were found in this image.")

           
           #cv2.imshow("contours", image_copy)

           cv2.imwrite(os.path.join(path1, rs + '.jpg'), image_copy)

           count = count + 1
           cv2.waitKey(0)
 
wb.save('Coordinates.xlsx')


count = 1
for filename in os.listdir(path1):
       img = cv2.imread(os.path.join(path1,filename))
       if img is not None: 
           rs = str(count)
           image = remove(img)
           cv2.imwrite(os.path.join(path2, rs + '.png'), image)
           count = count + 1


fname = "Coordinates.xlsx"
wb = openpyxl.load_workbook(fname)
ws = wb['Sheet1']

count = 1
for filename in os.listdir(folder):
       img1 = cv2.imread(os.path.join(folder,filename))
       img2 = cv2.imread(os.path.join(path2,filename))
       if img1 is not None: 
           rs = str(count)
           img1 = image_resize(img1, height = 600)
           x1 = ws.cell(1,count).value
           y1 = ws.cell(2,count).value
           width = ws.cell(3,count).value
           height = ws.cell(4,count).value

           imgx = np.array(img2)
           alpha_mask = imgx[:, :, :3] / 255.0
           img1 = img1[:, :, :3].copy()
           img2 = imgx[:, :, :3]
           alpha = alpha_mask[:height, :width, :np.newaxis]
           alpha_inv = 1.0 - alpha

           img1[y1:y1 + height, x1:x1+width, :] = (img2[:height, :width, :] * alpha) + (img1[y1:y1 + height, x1:x1+width, :] * alpha_inv)

           cv2.imwrite(os.path.join(path3, rs + '.png'), img1)
           count = count + 1

wb.save('Coordinates.xlsx')